# -*- coding: utf-8 -*-
"""
opencode.ai 使用历史（Usage History）爬取脚本 · 纯 HTTP 版

爬取 https://opencode.ai/workspace/{workspace}/usage 页面的"使用历史：近期 API 使用情况和成本"，
字段包括：输入(Input)、缓存读取(Cache Read)、输出(Output)、推理(Reasoning)、成本(Cost)、Session 等，并写入 Excel。

数据来源：SolidStart server function `usage.list`（POST /_server，带登录态 cookie）。
无需启动浏览器：直接复用 Playwright 登录后保存的 storage_state 登录态，逐页请求 API 即可，
速度远快于浏览器点击翻页。仅当没有缓存登录态 / 强制重新登录时才走 Playwright OAuth 登录。

用法:
    python scrape_usage.py                       # 爬取全部数据（默认仅最近 1 个月）
    python scrape_usage.py --months 0            # 全部数据
    python scrape_usage.py --months 3            # 最近 3 个月
    python scrape_usage.py --start 2026-08-01 --end 2026-08-06   # 指定时间段
    python scrape_usage.py --force-login --email x --password y  # 强制重新登录后爬取
    python scrape_usage.py --out result.xlsx
"""
import argparse
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import httpx
from openpyxl import Workbook

# ============ 配置 ============
BASE_URL = "https://opencode.ai"
WORKSPACE = None   # 运行时由 get_workspace() 从环境变量 / 本地配置获取
USAGE_URL = None   # 运行时按 WORKSPACE 拼接
# 代理直接从环境变量读取；未设置则直连（不使用代理）
PROXY = (os.environ.get("HTTPS_PROXY") or os.environ.get("HTTP_PROXY") or
         os.environ.get("ALL_PROXY") or "").strip()
OUT_DIR = Path(__file__).parent
LOCAL_TZ = timezone(timedelta(hours=8))  # 页面显示时区 +08:00
PAGE_SIZE = 50  # usage.list 每页条数

# 运行时配置缓存（含 workspace，被 .gitignore 忽略，不写入代码/提交）
CONFIG_FILE = OUT_DIR / "opencode_env.json"


def load_config() -> dict:
    """读取本地运行时配置（opencode_env.json）。"""
    if CONFIG_FILE.exists():
        try:
            data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {}


def get_workspace() -> str:
    """获取工作区 ID：优先环境变量（手动覆盖），其次本地缓存（登录时自动提取保存）。"""
    ws = os.environ.get("OPENCODE_WORKSPACE_ID", "").strip()
    if ws:
        return ws
    ws = (load_config().get("workspace") or "").strip()
    if ws:
        return ws
    return ""


def extract_workspace(url_or_text) -> str:
    """从 URL 或文本中提取 opencode.ai 工作区 ID（wrk_...），提取不到返回空串。"""
    if not url_or_text:
        return ""
    m = re.search(r"wrk_[A-Za-z0-9]+", str(url_or_text))
    return m.group(0) if m else ""


def save_workspace(ws: str) -> None:
    """把自动获取的工作区 ID 缓存到本地配置（opencode_env.json，被 .gitignore 忽略）。"""
    if not ws:
        return
    cfg = load_config()
    cfg["workspace"] = ws
    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

# usage.list server function 引用 ID（来自客户端 bundle: createServerReference(...)）
USAGE_LIST_ID = "bfd684bfc2e4eed05cd0b518f5e4eafd3f3376e3938abb9e536e7c03df831e5c"
# =============================


def state_file_for(method: str) -> Path:
    """登录态缓存文件按登录方式区分（github / google）。"""
    return OUT_DIR / f"storage_state_{method}.json"


def cookie_header(method: str) -> str:
    """从 storage_state 读取 opencode.ai 域的 cookie，拼成 Cookie header。"""
    state_file = state_file_for(method)
    if not state_file.exists():
        return ""
    try:
        state = json.loads(state_file.read_text(encoding="utf-8"))
    except Exception:
        return ""
    parts = []
    for c in state.get("cookies", []):
        dom = c.get("domain", "")
        if "opencode.ai" in dom:
            parts.append(f"{c['name']}={c['value']}")
    return "; ".join(parts)


def make_client(method: str) -> httpx.Client:
    """构造带登录态 cookie 与代理的 HTTP 客户端。"""
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": USAGE_URL or BASE_URL,
    }
    ck = cookie_header(method)
    if ck:
        headers["Cookie"] = ck
    if not PROXY:
        return httpx.Client(headers=headers, timeout=60)
    try:
        return httpx.Client(headers=headers, proxy=PROXY, timeout=60)
    except TypeError:
        return httpx.Client(headers=headers, proxies={"http://": PROXY, "https://": PROXY}, timeout=60)


# ============ seroval 请求/响应解析 ============
def seroval_body(page: int) -> dict:
    """构造 usage.list 的 seroval 参数 body：参数数组 [workspace, page]。"""
    return {
        "t": {"t": 9, "i": 0, "l": 2, "a": [
            {"t": 1, "s": WORKSPACE},
            {"t": 0, "s": page},
        ], "o": 0},
        "f": 31,
        "m": [],
    }


def extract_array_text(body: str) -> str | None:
    """从 _server 响应提取 usage 记录数组的 seroval 文本（$R[0]=[...]）。"""
    marker = "$R[0]=["
    start = body.find(marker)
    if start == -1:
        return None
    arr_start = start + len(marker) - 1  # 指向 '['
    end_marker = '])($R["server-fn'
    end = body.find(end_marker, arr_start)
    if end != -1:
        return body[arr_start:end + 1]
    # 兜底：括号匹配
    depth = 0
    i = arr_start
    while i < len(body):
        c = body[i]
        if c == '[':
            depth += 1
        elif c == ']':
            depth -= 1
            if depth == 0:
                return body[arr_start:i + 1]
        i += 1
    return body[arr_start:]


def seroval_to_json(text: str) -> str:
    """把 seroval JS 流文本转换为合法 JSON 文本（纯 Python，无需浏览器 eval）。"""
    t = text
    t = re.sub(r'new Date\("([^"]*)"\)', r'"\1"', t)  # new Date("...") → "..."
    t = re.sub(r'!0', 'true', t)
    t = re.sub(r'!1', 'false', t)
    t = re.sub(r'\$R\[\d+\]=', '', t)                 # 移除 $R[n]= 引用前缀
    t = re.sub(r'([{,]\s*)([A-Za-z_$][\w$]*)(\s*:)', r'\1"\2"\3', t)  # 对象 key 加引号
    return t


def parse_records(body: str) -> list:
    """从 _server 响应解析出 usage 记录列表。"""
    arr = extract_array_text(body)
    if not arr:
        return []
    try:
        data = json.loads(seroval_to_json(arr))
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    return [r for r in data if isinstance(r, dict)]


def fetch_usage_page(client: httpx.Client, page: int) -> list:
    """请求一页 usage 记录（POST /_server）。"""
    resp = client.post(
        f"{BASE_URL}/_server",
        headers={
            "Content-Type": "application/json",
            "X-Server-Id": USAGE_LIST_ID,
            "X-Server-Instance": f"server-fn:{page}",
        },
        json=seroval_body(page),
    )
    if resp.status_code == 401 or resp.status_code == 403:
        raise RuntimeError("登录态已失效，请提供账号密码强制重新登录（--force-login --email ... --password ...）")
    # opencode.ai 登录态失效时不返回 401，而是返回 HTTP 200 + seroval 302 重定向到
    # /auth/authorize（响应体形如 $R[0]=new Response(null,{headers:[["location","/auth/authorize"]],status:302,...})）。
    # 若不识别，会静默解析为空列表导致"爬空"，Excel 无法自动创建/填充。
    if "/auth/" in resp.text:
        raise RuntimeError("登录态已失效，请提供账号密码强制重新登录（--force-login --email ... --password ...）")
    if resp.status_code != 200:
        raise RuntimeError(f"请求 _server 失败: HTTP {resp.status_code} {resp.text[:200]}")
    return parse_records(resp.text)


def discover_workspace(method: str) -> str:
    """用缓存登录态请求 /workspace（跟随重定向），自动提取工作区 ID；失败返回空串。"""
    try:
        with make_client(method) as client:
            resp = client.get(f"{BASE_URL}/workspace", follow_redirects=True, timeout=30)
            ws = extract_workspace(str(resp.url))
            if not ws:
                ws = extract_workspace(resp.text)
            return ws
    except Exception:
        return ""


# ============ 登录（仅当需要时使用 Playwright 完成 OAuth） ============
def ensure_login(email=None, password=None, force_login=False, method="github") -> str:
    """按登录方式执行 OAuth 登录并保存 storage_state；登录完成后从 URL 自动提取并返回工作区 ID。"""
    from playwright.async_api import async_playwright  # 延迟导入：平时爬取不加载

    state_file = state_file_for(method)
    if force_login and state_file.exists():
        state_file.unlink()
    login_email = email or ""
    login_password = password or ""
    import asyncio

    async def _do_login():
        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=False,  # 可见模式：OAuth 流程与账号安全验证（验证码/设备码）对用户可见，遇验证码可手动完成
                proxy={"server": PROXY},
                args=["--disable-blink-features=AutomationControlled"],
            )
            ctx = await browser.new_context(
                viewport={"width": 1440, "height": 900}, locale="en-US",
                storage_state=str(state_file) if state_file.exists() else None,
            )
            page = await ctx.new_page()
            page.set_default_timeout(45000)
            # 打开 opencode.ai 首页开始 OAuth（不依赖 workspace；登录完成后自动提取 workspace）
            await page.goto(f"{BASE_URL}/", wait_until="domcontentloaded", timeout=60000)
            await page.wait_for_timeout(3000)
            body = await page.evaluate("document.body.innerText")

            async def wait_back_to_opencode(timeout_ms=180000):
                """等待登录/授权完成并回到 opencode.ai；若出现账号安全验证码，提示用户手动完成。"""
                import time as _t
                deadline = _t.time() + timeout_ms / 1000
                while _t.time() < deadline:
                    cur = page.url
                    try:
                        b = await page.evaluate("document.body ? document.body.innerText : ''")
                    except Exception:
                        b = ""
                    # 已回到 opencode.ai 且不在授权/登录页 → 完成
                    if "opencode.ai" in cur and "authorize" not in cur.lower() and "/auth/" not in cur:
                        return True
                    # 检测到账号安全验证 / 设备码 / 二次验证页 → 提示用户在浏览器中完成
                    low = cur.lower()
                    if any(k in low for k in ("two-factor", "login/device", "verification",
                                              "challenge", "enter the code")):
                        print("检测到账号安全验证，请在浏览器窗口中完成验证码 / 设备码授权…")
                    await page.wait_for_timeout(1500)
                return False

            if method == "github" and "Continue with GitHub" in body:
                await page.click("text=Continue with GitHub", timeout=15000)
                await page.wait_for_timeout(4000)
                # 三层：仅当 GitHub 登录表单存在（关联账号会话失效）才填账号密码；否则 provider 已登录自动授权
                if await page.locator('#login_field').count() > 0:
                    if not login_email or not login_password:
                        raise RuntimeError("关联账号(GitHub)会话已失效，需要账号密码登录（请在设置页填写账号密码）")
                    for _ in range(3):
                        try:
                            await page.fill("#login_field", login_email, timeout=5000)
                            await page.click("input[type=submit]", timeout=5000)
                            await page.wait_for_timeout(2500)
                        except Exception:
                            break
                        if "github.com" not in page.url or "login" not in page.url:
                            break
                    try:
                        await page.fill("#password", login_password, timeout=10000)
                        await page.click("input[type=submit]", timeout=10000)
                        await page.wait_for_timeout(5000)
                    except Exception:
                        pass
                # 等待回到 opencode.ai；若出现 2FA / 设备码，由用户在浏览器中手动完成验证
                if not await wait_back_to_opencode():
                    raise RuntimeError("登录超时：可能遇到账号验证码需要手动完成，或授权未在浏览器中完成")
                try:
                    b = await page.evaluate("document.body.innerText")
                    if "Authorize" in b and "opencode" in b.lower():
                        await page.click("button:has-text('Authorize')", timeout=8000)
                        await page.wait_for_timeout(2000)
                except Exception:
                    pass
            elif method == "google" and "Continue with Google" in body:
                await page.click("text=Continue with Google", timeout=15000)
                await page.wait_for_timeout(6000)
                # 三层：仅当 Google 邮箱输入框存在（关联账号会话失效）才填账号密码；否则 provider 已登录自动授权
                if await page.locator('input[type="email"]').count() > 0:
                    if not login_email or not login_password:
                        raise RuntimeError("关联账号(Google)会话已失效，需要账号密码登录（请在设置页填写账号密码）")
                    try:
                        await page.fill('input[type="email"]', login_email, timeout=6000)
                        await page.click("#identifierNext", timeout=6000)
                        await page.wait_for_timeout(4000)
                    except Exception:
                        pass
                    try:
                        await page.fill('input[type="password"]', login_password, timeout=10000)
                        await page.click("#passwordNext", timeout=10000)
                        await page.wait_for_timeout(6000)
                    except Exception:
                        pass
                # 等待回到 opencode.ai；若出现验证码 / 二次验证，由用户在浏览器中手动完成
                if not await wait_back_to_opencode():
                    raise RuntimeError("登录超时：可能遇到账号验证码需要手动完成，或授权未在浏览器中完成")
            else:
                raise RuntimeError(f"登录方式 {method} 不受支持或页面未出现对应登录入口")
            await ctx.storage_state(path=str(state_file))
            # 登录完成：从当前 URL 提取工作区 ID；提取不到则访问 /workspace 等待重定向提取
            ws = extract_workspace(page.url)
            if not ws:
                try:
                    await page.goto(f"{BASE_URL}/workspace", wait_until="domcontentloaded", timeout=30000)
                    await page.wait_for_timeout(2500)
                    ws = extract_workspace(page.url)
                except Exception:
                    pass
            await browser.close()
            return ws

    return asyncio.run(_do_login())


# ============ 数据格式化 ============
def fmt_cost(cost) -> float:
    """cost 原始值转美元：cost / 1e8。"""
    try:
        return round(float(cost) / 1e8, 8)
    except (TypeError, ValueError):
        return cost


def fmt_cost_display(cost) -> str:
    """页面展示格式：Go ($0.0002)。"""
    try:
        return f"Go (${float(cost) / 1e8:.4f})"
    except (TypeError, ValueError):
        return str(cost)


def to_local(iso_str: str) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.astimezone(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return iso_str


def record_to_row(r: dict) -> dict:
    return {
        "日期时间": to_local(r.get("timeCreated") or ""),
        "月份": (r.get("timeCreated") or "")[:7],
        "模型": r.get("model", ""),
        "提供方": r.get("provider", ""),
        "输入 Input": r.get("inputTokens", 0),
        "缓存读取 Cache Read": r.get("cacheReadTokens", 0),
        "输入合计": (r.get("inputTokens", 0) or 0) + (r.get("cacheReadTokens", 0) or 0),
        "输出 Output": r.get("outputTokens", 0),
        "推理 Reasoning": r.get("reasoningTokens", 0),
        "输出合计": (r.get("outputTokens", 0) or 0),
        "成本($)": fmt_cost(r.get("cost", 0)),
        "成本显示": fmt_cost_display(r.get("cost", 0)),
        "Session": r.get("sessionID", "") or "",
        "Key ID": r.get("keyID", ""),
        "计划": (r.get("enrichment") or {}).get("plan", "") if isinstance(r.get("enrichment"), dict) else "",
        "记录ID": r.get("id", ""),
    }


def add_records(records: list, all_records: list, seen: set) -> int:
    """按 id 去重后加入记录列表，返回新增条数。"""
    added = 0
    for r in records:
        rid = r.get("id")
        if rid and rid not in seen:
            seen.add(rid)
            all_records.append(r)
            added += 1
    return added


# ============ 时间段过滤 ============
def to_local_dt(iso: str):
    """UTC ISO 时间 → 本地（+08:00）naive datetime；解析失败返回 None。"""
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).astimezone(LOCAL_TZ).replace(tzinfo=None)
    except Exception:
        return None


def read_latest_time(out: str):
    """读取 Excel 第一行数据的本地日期时间（爬虫按时间倒序写入，第一行即最新）。"""
    path = OUT_DIR / out
    if not path.exists():
        return None
    wb = None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return None
        idx = next((i for i, h in enumerate(header) if h == "日期时间"), None)
        if idx is None:
            return None
        for row in rows:
            if row and row[idx]:
                try:
                    return datetime.fromisoformat(str(row[idx]))
                except Exception:
                    pass
            break
    except Exception:
        return None
    finally:
        # 无论提前 return 还是异常，都必须释放 workbook 句柄（否则 Windows 下文件被锁）
        if wb is not None:
            wb.close()
    return None


def local_to_utc_iso(local_str) -> str:
    """Excel 本地时间字符串 → UTC ISO（供记录排序/比较）。"""
    try:
        s = str(local_str).strip()
        dt = datetime.fromisoformat(s.replace(" ", "T") if "T" not in s else s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=LOCAL_TZ)
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    except Exception:
        return ""


def read_excel_records(out: str) -> list:
    """读取旧 Excel 全部记录，转回原始记录格式（用于增量合并）。"""
    path = OUT_DIR / out
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h) for h in next(rows, ())]
        if not header:
            wb.close()
            return []

        def col(name):
            return header.index(name) if name in header else -1

        c_time, c_model, c_provider = col("日期时间"), col("模型"), col("提供方")
        c_input, c_cache = col("输入 Input"), col("缓存读取 Cache Read")
        c_output, c_reason = col("输出 Output"), col("推理 Reasoning")
        c_cost, c_key, c_plan, c_session, c_id = col("成本($)"), col("Key ID"), col("计划"), col("Session"), col("记录ID")
        records = []
        for row in rows:
            if not row or c_id < 0 or not row[c_id]:
                continue
            rec = {"id": str(row[c_id])}
            if c_time >= 0 and row[c_time]:
                rec["timeCreated"] = local_to_utc_iso(row[c_time])
            if c_model >= 0:
                rec["model"] = row[c_model] or ""
            if c_provider >= 0:
                rec["provider"] = row[c_provider] or ""
            if c_input >= 0:
                rec["inputTokens"] = row[c_input] or 0
            if c_cache >= 0:
                rec["cacheReadTokens"] = row[c_cache] or 0
            if c_output >= 0:
                rec["outputTokens"] = row[c_output] or 0
            if c_reason >= 0:
                rec["reasoningTokens"] = row[c_reason] or 0
            if c_cost >= 0 and row[c_cost] is not None:
                try:
                    rec["cost"] = round(float(row[c_cost]) * 1e8)
                except Exception:
                    rec["cost"] = 0
            if c_key >= 0:
                rec["keyID"] = row[c_key] or ""
            if c_session >= 0:
                rec["sessionID"] = row[c_session] or ""
            if c_plan >= 0 and row[c_plan]:
                rec["enrichment"] = {"plan": str(row[c_plan])}
            records.append(rec)
        wb.close()
        return records
    except Exception:
        return []


def filter_records(records: list, months: int, start: str, end: str) -> list:
    """按月份数 / 指定时间段过滤（基于本地时间）。"""
    if not months and not start and not end:
        return records

    if months and months > 0:
        cutoff = (datetime.now(LOCAL_TZ).replace(tzinfo=None) - timedelta(days=365.25 / 12 * months))
    else:
        cutoff = None
    start_dt = datetime.fromisoformat(start) if start else None
    end_dt = datetime.fromisoformat(end + "T23:59:59") if end else None

    filtered = []
    for r in records:
        t = to_local_dt(r.get("timeCreated") or "")
        if t is None:
            filtered.append(r)
            continue
        if cutoff and t < cutoff:
            continue
        if start_dt and t < start_dt:
            continue
        if end_dt and t > end_dt:
            continue
        filtered.append(r)
    return filtered


def save_excel(records: list, out: str) -> Path:
    rows = [record_to_row(r) for r in records]
    wb = Workbook()
    ws = wb.active
    ws.title = "Usage History"
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    for col_idx, h in enumerate(headers, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(14, len(h) * 2 + 4)
    out_path = OUT_DIR / out
    wb.save(str(out_path))
    return out_path


# ============ 主流程 ============
def main() -> None:
    parser = argparse.ArgumentParser(description="爬取 opencode.ai 使用历史（纯 HTTP）")
    parser.add_argument("--months", type=int, default=1,
                        help="保留最近几个月（默认1=最近1个月；0=全部数据）")
    parser.add_argument("--out", type=str, default="opencode_usage_history.xlsx", help="输出 Excel 文件名")
    parser.add_argument("--email", type=str, default="", help="登录账号（留空用缓存登录态）")
    parser.add_argument("--password", type=str, default="", help="登录密码")
    parser.add_argument("--login-method", type=str, default="github", choices=["github", "google"],
                        help="登录方式（github / google）")
    parser.add_argument("--workspace", type=str, default="",
                        help="工作区 ID（wrk_...）；留空则登录后自动获取 / 复用本地缓存")
    parser.add_argument("--force-login", action="store_true", help="强制重新登录（忽略缓存登录态）")
    parser.add_argument("--start", type=str, default="", help="下载时间段起点 YYYY-MM-DD（本地时间）")
    parser.add_argument("--end", type=str, default="", help="下载时间段终点 YYYY-MM-DD（本地时间，含当天）")
    parser.add_argument("--incremental", action="store_true",
                        help="增量更新：仅爬取 Excel 最新记录之后的新数据（遇到更早记录即停止）")
    args = parser.parse_args()

    # 工作区 ID 解析：显式 --workspace → 环境变量/本地缓存 → 登录时自动提取 → 缓存登录态自动发现
    global WORKSPACE, USAGE_URL
    WORKSPACE = (args.workspace or get_workspace()).strip()

    # 登录：无缓存登录态或强制重新登录时走三层恢复（缓存→关联账号→账号密码+验证码），
    # 登录完成后自动从浏览器 URL 提取工作区 ID（无需用户手动配置）
    if not state_file_for(args.login_method).exists() or args.force_login:
        print("需要登录 / 重新授权 opencode.ai ...")
        ws = ensure_login(args.email, args.password, args.force_login, args.login_method)
        print("登录态已保存")
        if not WORKSPACE:
            WORKSPACE = ws or ""
    # 未获取到且已有缓存登录态：请求 /workspace 跟随重定向自动提取
    if not WORKSPACE and state_file_for(args.login_method).exists():
        print("尝试从缓存登录态自动发现工作区 ...")
        WORKSPACE = discover_workspace(args.login_method)
    if not WORKSPACE:
        raise RuntimeError(
            "无法自动获取工作区 ID（wrk_...）。请先登录后再运行，或用 --workspace 手动指定。")
    save_workspace(WORKSPACE)
    USAGE_URL = f"{BASE_URL}/workspace/{WORKSPACE}/usage"

    # 增量起点：Excel 最新一条记录的本地时间（数据按时间倒序写入，第一行即最新）
    incremental_start = read_latest_time(args.out) if args.incremental else None
    if incremental_start is not None:
        print(f"增量起点（Excel 最新）: {incremental_start}")

    all_records = []
    seen = set()
    reached_start = False
    auth_attempts = 0
    net_retries = 0
    while True:
        try:
            with make_client(args.login_method) as client:
                page = 0
                while True:
                    records = fetch_usage_page(client, page)
                    # 增量模式：数据按时间倒序，遇到 ≤ 增量起点的记录即停止（其后均为旧数据）
                    if incremental_start is not None:
                        new_records = []
                        for r in records:
                            t = to_local_dt(r.get("timeCreated") or "")
                            if t is not None and t <= incremental_start:
                                reached_start = True
                                break
                            new_records.append(r)
                        added = add_records(new_records, all_records, seen)
                        suffix = "（到达增量起点，停止）" if reached_start else ""
                        print(f"  第{page + 1}页: {added} 条{suffix}")
                        if reached_start or len(records) < PAGE_SIZE:
                            break
                    else:
                        added = add_records(records, all_records, seen)
                        print(f"  第{page + 1}页: {added} 条")
                        if len(records) < PAGE_SIZE:
                            break
                    page += 1
            break
        except RuntimeError as e:
            # opencode 登录态失效 → 三层恢复（关联账号自动授权 → 账号密码），然后重试
            if "登录态已失效" in str(e) and auth_attempts < 2:
                auth_attempts += 1
                print("opencode 登录态失效，尝试重新授权（关联账号 / 账号密码）...")
                try:
                    ensure_login(args.email, args.password, False, args.login_method)
                except RuntimeError as e2:
                    print(f"[fatal] {e2}")
                    sys.exit(1)
                continue
            print(f"[fatal] {e}")
            sys.exit(1)
        except Exception as e:
            # 网络 / SSL 偶发错误（如代理瞬时中断、连接被重置）→ 短暂重试，避免一次抖动导致整个更新失败
            if net_retries < 3:
                net_retries += 1
                print(f"[retry] 网络异常: {type(e).__name__}: {e}，第 {net_retries}/3 次重试...")
                time.sleep(2)
                continue
            print(f"[fatal] 网络异常持续: {e}")
            sys.exit(1)

    # 按时间倒序
    all_records.sort(key=lambda r: r.get("timeCreated") or "", reverse=True)
    print(f"共 {page + 1} 页，去重后 {len(all_records)} 条")

    # 过滤（月份 / 时间段）
    filtered = filter_records(all_records, args.months, args.start, args.end)
    if args.start or args.end:
        print(f"时间段过滤后记录: {len(filtered)}")
    elif args.months and args.months > 0:
        print(f"最近 {args.months} 个月过滤后记录: {len(filtered)}")

    # 增量模式：与旧 Excel 数据合并去重，保留完整历史（只追加新记录，避免覆盖丢失）
    if incremental_start is not None:
        merged = {r["id"]: r for r in read_excel_records(args.out) if r.get("id")}
        for r in filtered:
            if r.get("id"):
                merged[r["id"]] = r
        filtered = list(merged.values())
        filtered.sort(key=lambda r: r.get("timeCreated") or "", reverse=True)
        print(f"合并旧 Excel 后总记录: {len(filtered)}")

    print(f"总计记录: {len(filtered)}")

    out_path = save_excel(filtered, args.out)
    print(f"Excel 已保存: {out_path}")


if __name__ == "__main__":
    main()
